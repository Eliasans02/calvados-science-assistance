"""Final report composition and persistence."""

from __future__ import annotations

import json
import uuid
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from src import config
from src.data.repository import BackendRepository


class ReportService:
    def __init__(self, repository: BackendRepository) -> None:
        self._repository = repository

    def compose(self, user_id: str, file_id: str, context: Optional[dict[str, Any]] = None) -> dict[str, Any]:
        file_record = self._repository.get_file(file_id=file_id, user_id=user_id)
        if not file_record:
            raise ValueError("File not found for this user")

        agent_results = self._repository.list_agent_results(user_id=user_id, file_id=file_id)
        payload = {
            "file": {
                "id": file_record["id"],
                "filename": file_record["filename"],
                "uploaded_at": file_record["uploaded_at"],
                "warning": file_record["warning"],
            },
            "agent_results": agent_results,
            "context": context or {},
        }
        payload["template_scores"] = self._build_template_scores(payload)

        markdown_report = self._build_markdown(payload)
        report_id = str(uuid.uuid4())
        md_path = config.REPORTS_DATA_DIR / f"{report_id}.md"
        json_path = config.REPORTS_DATA_DIR / f"{report_id}.json"
        md_path.write_text(markdown_report, encoding="utf-8")
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        self._repository.save_report(
            user_id=user_id,
            file_id=file_id,
            report_format="markdown",
            report_path=str(md_path),
            report_payload=payload,
        )

        return {
            "agent": "report_agent",
            "report_id": report_id,
            "markdown_path": str(md_path),
            "json_path": str(json_path),
            "summary": {
                "agent_steps": len(agent_results),
                "file_id": file_id,
            },
            "report": payload,
        }

    def render_template_xlsx(self, report_payload: dict[str, Any], output_path: Path) -> Path:
        template_path = config.REPORT_TEMPLATE_XLSX_PATH
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        workbook = load_workbook(template_path)
        worksheet = workbook["Оценка ТЗ"] if "Оценка ТЗ" in workbook.sheetnames else workbook.active
        scores = report_payload.get("template_scores") or self._build_template_scores(report_payload)
        context = report_payload.get("context") or {}
        file_data = report_payload.get("file") or {}

        worksheet["A2"] = 1
        worksheet["B2"] = file_data.get("filename", "")
        worksheet["C2"] = context.get("organization", "")
        worksheet["D2"] = context.get("expert", "Calvados AI")
        worksheet["E2"] = scores["strategic_relevance"]
        worksheet["F2"] = scores["goal_and_tasks"]
        worksheet["G2"] = scores["scientific_novelty"]
        worksheet["H2"] = scores["practical_applicability"]
        worksheet["I2"] = scores["expected_results"]
        worksheet["J2"] = scores["socioeconomic_effect"]
        worksheet["K2"] = scores["feasibility"]
        text_blocks = self._build_template_text_blocks(report_payload)

        text_columns = {
            "M": "Комментарий: стратегическая релевантность",
            "N": "Комментарий: цель и задачи",
            "O": "Комментарий: научная новизна",
            "P": "Комментарий: практическая применимость",
            "Q": "Комментарий: ожидаемые результаты",
            "R": "Комментарий: соц-экономический эффект",
            "S": "Комментарий: реализуемость",
            "T": "AI резюме",
            "U": "Приоритетные действия",
            "V": "Quick wins",
            "W": "Итоговая рекомендация",
            "X": "Прозрачность анализа",
        }
        for column, header in text_columns.items():
            worksheet[f"{column}1"] = header
            worksheet[f"{column}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet.column_dimensions[column].width = 42

        worksheet["M2"] = text_blocks["strategic_comment"]
        worksheet["N2"] = text_blocks["goal_comment"]
        worksheet["O2"] = text_blocks["novelty_comment"]
        worksheet["P2"] = text_blocks["practical_comment"]
        worksheet["Q2"] = text_blocks["expected_results_comment"]
        worksheet["R2"] = text_blocks["socio_comment"]
        worksheet["S2"] = text_blocks["feasibility_comment"]
        worksheet["T2"] = text_blocks["ai_summary"]
        worksheet["U2"] = text_blocks["priority_actions"]
        worksheet["V2"] = text_blocks["quick_wins"]
        worksheet["W2"] = text_blocks["final_recommendation"]
        worksheet["X2"] = text_blocks["transparency_note"]
        for column in text_columns.keys():
            worksheet[f"{column}2"].alignment = Alignment(vertical="top", wrap_text=True)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _build_markdown(self, payload: dict[str, Any]) -> str:
        by_agent = self._latest_agent_outputs(payload.get("agent_results", []))
        recommendation = by_agent.get("recommendation", {})
        ai_enhancement = recommendation.get("ai_enhancement") or {}
        transparency = recommendation.get("transparency") or {}
        score_output = by_agent.get("scoring", {})
        lines = [
            "# Итоговый отчёт анализа ТЗ",
            "",
            f"- **Файл:** {payload['file']['filename']}",
            f"- **File ID:** {payload['file']['id']}",
            f"- **Дата загрузки:** {payload['file']['uploaded_at']}",
            "",
            "## Оценка по шаблону",
            f"- Стратегическая релевантность: **{payload['template_scores']['strategic_relevance']}/10**",
            f"- Цель и задачи: **{payload['template_scores']['goal_and_tasks']}/10**",
            f"- Научная новизна: **{payload['template_scores']['scientific_novelty']}/10**",
            f"- Практическая применимость: **{payload['template_scores']['practical_applicability']}/10**",
            f"- Ожидаемые результаты: **{payload['template_scores']['expected_results']}/10**",
            f"- Соц-экономический эффект: **{payload['template_scores']['socioeconomic_effect']}/10**",
            f"- Реализуемость: **{payload['template_scores']['feasibility']}/10**",
            f"- Интегральный скоринг: **{score_output.get('score', '-')}/{score_output.get('max_score', 100)}**",
            "",
            "## Рекомендации к доработке",
        ]
        for rec in (recommendation.get("recommendations") or [])[:8]:
            lines.append(f"- **{rec.get('priority', 'Medium')}** · {rec.get('text', '')}")
        if ai_enhancement.get("status") == "ready":
            lines.extend(
                [
                    "",
                    "### AI-обогащение рекомендаций",
                    f"- Резюме: {ai_enhancement.get('executive_summary', '')}",
                ]
            )
            for action in (ai_enhancement.get("priority_actions") or [])[:5]:
                lines.append(
                    f"- **{action.get('priority', 'Medium')}** [{action.get('section', 'раздел')}] "
                    f"{action.get('problem', '')} → {action.get('action', '')}"
                )
            if ai_enhancement.get("final_recommendation"):
                lines.append(f"- Итог: {ai_enhancement.get('final_recommendation')}")
        elif recommendation:
            lines.append(
                f"- AI-обогащение: {ai_enhancement.get('summary') or ai_enhancement.get('reason', 'недоступно')}"
            )
        if transparency:
            quality = (transparency.get("text_quality") or {}).get("quality_flag", "unknown")
            lines.extend(
                [
                    "",
                    "## Прозрачность анализа",
                    f"- Длина извлечённого текста: **{transparency.get('text_length', 0)}** символов",
                    f"- Качество текста: **{quality}**",
                    f"- Уверенность автоанализа: **{transparency.get('analysis_confidence', 'medium')}**",
                ]
            )
            for note in transparency.get("notes") or []:
                lines.append(f"- {note}")
        lines.extend(
            [
                "",
            "## Промежуточные результаты агентов",
            ]
        )
        if not payload["agent_results"]:
            lines.append("- Агентные результаты пока отсутствуют.")
        for result in payload["agent_results"]:
            lines.extend(
                [
                    "",
                    f"### {result['agent_name']}",
                    f"- Время: {result['created_at']}",
                    "```json",
                    json.dumps(result["output_json"], ensure_ascii=False, indent=2),
                    "```",
                ]
            )
        return "\n".join(lines)

    def _build_template_scores(self, payload: dict[str, Any]) -> dict[str, float]:
        by_agent = self._latest_agent_outputs(payload.get("agent_results", []))

        requirement = by_agent.get("requirement-analysis", {})
        text_analysis = by_agent.get("text-analysis", {})
        compliance = by_agent.get("compliance", {})
        scoring = by_agent.get("scoring", {})

        missing_sections = set(requirement.get("missing_sections") or [])
        missing_kpi = bool(requirement.get("missing_kpi"))
        text_issues = int((text_analysis.get("summary") or {}).get("total_issues", 0))
        passed = set(compliance.get("passed_items") or [])
        failed = set(compliance.get("missing_items") or [])
        rules_total = max(1, len(passed) + len(failed))
        compliance_ratio = len(passed) / rules_total
        overall_score_100 = float(scoring.get("score", 0)) if scoring else 0.0

        strategic_relevance = 4.0 + compliance_ratio * 6.0
        if "strategic_points" in passed:
            strategic_relevance += 1.0
        if "strategic_points" in failed:
            strategic_relevance -= 1.0

        goal_and_tasks = 10.0
        if "2.1_goal" in missing_sections:
            goal_and_tasks -= 3.5
        if "2.2_tasks" in missing_sections:
            goal_and_tasks -= 3.5
        if missing_kpi:
            goal_and_tasks -= 1.0

        scientific_novelty = 10.0 - (text_issues * 0.7)

        practical_applicability = 5.0 + (compliance_ratio * 3.0)
        if "consumers" in passed:
            practical_applicability += 1.0
        if "budget_by_year" in passed:
            practical_applicability += 1.0

        expected_results = 10.0
        if "4.1_direct_results" in missing_sections:
            expected_results -= 3.5
        if "4.2_final_result" in missing_sections:
            expected_results -= 3.5
        if missing_kpi:
            expected_results -= 2.0
        if "publication_targets" in failed:
            expected_results -= 1.0
        if "patent_targets" in failed:
            expected_results -= 1.0

        socioeconomic_effect = 5.0
        if "consumers" in passed:
            socioeconomic_effect += 1.5
        if "strategic_points" in passed:
            socioeconomic_effect += 1.0
        if not missing_kpi:
            socioeconomic_effect += 2.0
        socioeconomic_effect += compliance_ratio * 0.5

        feasibility = 10.0 - len(missing_sections) * 0.8 - text_issues * 0.2 - len(failed) * 0.3
        if overall_score_100 > 0:
            feasibility = (feasibility + overall_score_100 / 10.0) / 2.0

        return {
            "strategic_relevance": self._clamp_10(strategic_relevance),
            "goal_and_tasks": self._clamp_10(goal_and_tasks),
            "scientific_novelty": self._clamp_10(scientific_novelty),
            "practical_applicability": self._clamp_10(practical_applicability),
            "expected_results": self._clamp_10(expected_results),
            "socioeconomic_effect": self._clamp_10(socioeconomic_effect),
            "feasibility": self._clamp_10(feasibility),
        }

    @staticmethod
    def _clamp_10(value: float) -> float:
        return round(max(0.0, min(10.0, float(value))), 2)

    @staticmethod
    def _latest_agent_outputs(agent_results: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        by_agent: dict[str, dict[str, Any]] = {}
        for item in agent_results:
            agent_name = str(item.get("agent_name", "")).strip()
            if not agent_name:
                continue
            by_agent[agent_name] = item.get("output_json") or {}
        return by_agent

    def _build_template_text_blocks(self, payload: dict[str, Any]) -> dict[str, str]:
        by_agent = self._latest_agent_outputs(payload.get("agent_results", []))
        requirement = by_agent.get("requirement-analysis", {})
        text_analysis = by_agent.get("text-analysis", {})
        compliance = by_agent.get("compliance", {})
        recommendation = by_agent.get("recommendation", {})
        scoring = by_agent.get("scoring", {})
        transparency = recommendation.get("transparency") or {}

        missing_sections = set(requirement.get("missing_sections") or [])
        missing_kpi = bool(requirement.get("missing_kpi"))
        passed = set(compliance.get("passed_items") or [])
        failed = set(compliance.get("missing_items") or [])
        text_summary = text_analysis.get("summary") or {}
        recommendations = recommendation.get("recommendations") or []
        ai = recommendation.get("ai_enhancement") or {}

        top_recs = [item.get("text", "").strip() for item in recommendations[:3] if item.get("text")]
        actions = ai.get("priority_actions") or []
        quick_wins = ai.get("quick_wins") or []

        strategic_comment = (
            "Есть явная связка со стратегическими документами."
            if "strategic_points" in passed
            else "Не хватает чёткой привязки к стратегическим/программным документам."
        )
        if top_recs:
            strategic_comment += f" Рекомендация: {top_recs[0]}"

        goal_parts: list[str] = []
        if "2.1_goal" in missing_sections:
            goal_parts.append("не раскрыта цель программы (2.1)")
        if "2.2_tasks" in missing_sections:
            goal_parts.append("не раскрыты задачи программы (2.2)")
        if missing_kpi:
            goal_parts.append("нет измеримых KPI")
        goal_comment = "Цель и задачи описаны удовлетворительно." if not goal_parts else " ; ".join(goal_parts).capitalize() + "."

        novelty_issues = int(text_summary.get("vague_formulations", 0)) + int(text_summary.get("logical_conflicts", 0))
        novelty_comment = (
            "Формулировки достаточно конкретные, явных логических конфликтов мало."
            if novelty_issues == 0
            else f"Найдены проблемные формулировки/конфликты: {novelty_issues}. Требуется уточнение научного вклада."
        )

        practical_parts: list[str] = []
        if "consumers" not in passed:
            practical_parts.append("не обозначены целевые потребители")
        if "budget_by_year" not in passed:
            practical_parts.append("не детализирован бюджет по годам")
        practical_comment = (
            "Практическая применимость подтверждена целевыми потребителями и планом реализации."
            if not practical_parts
            else " ; ".join(practical_parts).capitalize() + "."
        )

        expected_parts: list[str] = []
        if "4.1_direct_results" in missing_sections:
            expected_parts.append("нет блока прямых результатов (4.1)")
        if "4.2_final_result" in missing_sections:
            expected_parts.append("нет конечного результата (4.2)")
        if "publication_targets" in failed:
            expected_parts.append("не указаны публикационные KPI")
        if "patent_targets" in failed:
            expected_parts.append("не указаны патентные KPI")
        expected_results_comment = (
            "Ожидаемые результаты сформулированы и поддаются измерению."
            if not expected_parts
            else " ; ".join(expected_parts).capitalize() + "."
        )

        socio_comment = (
            "Соц-экономический эффект описан и связан с практическим внедрением."
            if ("consumers" in passed and not missing_kpi)
            else "Нужно усилить описание социально-экономического эффекта и метрик внедрения."
        )

        score = float(scoring.get("score", 0))
        feasibility_comment = (
            f"Интегральный скоринг: {score}/100. Реализуемость высокая."
            if score >= 75
            else f"Интегральный скоринг: {score}/100. Требуется доработка критичных разделов."
        )

        ai_summary = (
            str(ai.get("executive_summary", "")).strip()
            or str(ai.get("summary", "")).strip()
            or "AI-сводка недоступна, использованы rule-based рекомендации."
        )

        if actions:
            priority_actions = "\n".join(
                f"- [{a.get('priority', 'Medium')}] {a.get('section', 'раздел')}: {a.get('action', '')}"
                for a in actions[:5]
            )
        elif top_recs:
            priority_actions = "\n".join(f"- {item}" for item in top_recs)
        else:
            priority_actions = "Критичных действий не выявлено."

        if quick_wins:
            quick_wins_text = "\n".join(f"- {item}" for item in quick_wins[:5])
        else:
            quick_wins_text = "\n".join(f"- {item}" for item in top_recs[:2]) if top_recs else "Быстрых улучшений не предложено."

        final_recommendation = (
            str(ai.get("final_recommendation", "")).strip()
            or "Сначала закрыть отсутствующие обязательные разделы и KPI, затем обновить формулировки под измеримые критерии."
        )
        transparency_note_parts = [
            f"Качество текста: {(transparency.get('text_quality') or {}).get('quality_flag', 'unknown')}",
            f"Уверенность: {transparency.get('analysis_confidence', 'medium')}",
            f"Длина текста: {transparency.get('text_length', 0)} симв.",
        ]
        transparency_notes = transparency.get("notes") or []
        if transparency_notes:
            transparency_note_parts.append("; ".join(str(item) for item in transparency_notes[:2]))
        transparency_note = " | ".join(transparency_note_parts)

        return {
            "strategic_comment": strategic_comment,
            "goal_comment": goal_comment,
            "novelty_comment": novelty_comment,
            "practical_comment": practical_comment,
            "expected_results_comment": expected_results_comment,
            "socio_comment": socio_comment,
            "feasibility_comment": feasibility_comment,
            "ai_summary": ai_summary,
            "priority_actions": priority_actions,
            "quick_wins": quick_wins_text,
            "final_recommendation": final_recommendation,
            "transparency_note": transparency_note,
        }
